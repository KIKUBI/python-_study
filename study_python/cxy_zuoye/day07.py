# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   RONGHUA
# FileName:      day07.py
# Author:       陈啸宇
# Datetime:     2024/8/7 15:54
# Description:
# ---------------------------------------------------------------------------
import faker
import openpyxl
class cxy_txt:
    def __init__(self,path_txt):
        self.path_txt = path_txt
    def creat_txt(self,n):
        with open(self.path_txt, "a", encoding="utf-8") as f:
            for i in range(1, n):
                f.write("############2222222222222\n")
                f.write('122222222222222\n')
    def selct_txt(self,x):
        with open(self.path_txt, "r", encoding="utf-8") as c:
            self.li = c.readlines()
            self.ls = []
            for i in range(len(self.li)):
                if self.li[i][0] != x:
                    self.ls.append(self.li[i])
            return self.ls
    def write_txt(self,ls):
        with open(self.path_txt, "w", encoding="utf-8") as x:
            x.write(''.join(ls))

cxytxt = cxy_txt("day08.txt")
cxytxt.creat_txt(10)
r = cxytxt.selct_txt('#') #去掉#开头的行
cxytxt.write_txt(r)
###################################################
class cxy_he:
    date=faker.Faker(locale="zh_CN")
    def __init__(self):
        pass
    def creat_cxyexcel(self,path,name,ends):
        self.w = openpyxl.Workbook()
        self.ws = self.w.create_sheet(name,0)
        self.ws.append(['编号',"姓名", "地址", "电话"])
        for i in range(1, ends):
            self.ws.append([i,self.date.name_male(),self.date.address(),self.date.phone_number()])
            self.w.save(path)
        return self.ws
    def copy_cxyexcel1(self,excel,path,name,first,end):
        self.w1 = openpyxl.Workbook()
        self.ws1 = self.w1.create_sheet(name,0)
        for j in excel.iter_rows(min_row=first,max_row=end,values_only=True):
            self.ws1.append(list(j))
            self.w1.save(path)
        return self.ws1
    def copy_cxyexcel2(self,excel,excel1,path,name,first1,end1,first2,end2):
        self.w2 = openpyxl.Workbook()
        self.ws2 = self.w2.create_sheet(name,0)
        for j in excel.iter_rows(min_row=first1,max_row=end1,values_only=True):
            self.ws2.append(list(j))
        for z in excel1.iter_rows(min_row=first2,max_row=end2,values_only=True):
            self.ws2.append(list(z))
            self.w2.save(path)
        return self.ws2
    def add_cxyexcel(self,excel2,path,first,end,sum):
        for d in range(first,end):
            excel2[f'A{d}'].value+=sum
        self.w2.save(path)
zzz = cxy_he()
s = zzz.creat_cxyexcel("cxy.xlsx","cxy",11) #创建一个excel为cxy
s1 = zzz.copy_cxyexcel1(s,"cxy1.xlsx","cxy",1,11)                         #复制excel为cxy1
s2 = zzz.copy_cxyexcel2(s,s1,"cxy2.xlsx","cxy",1,11,2,11)                 #复制excel为cxy2
zzz.add_cxyexcel(s2,"cxy2.xlsx",12,22,10)                                 #在cxy2.xlsx中A列12-22行加10

#########################################################################
class cxy_add:
    def __init__(self,path,first,end,d):
        self.path = path
        self.first = first
        self.end = end
        self.d = d
        self.result1 = 0
        self.result2 = 0
        self.jsum = []
        self.osum = []
    def creat_cxyexcel(self):
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb["Sheet1"]
        self.lsz = []
        for self.row in self.ws.iter_rows(min_row=self.first,max_row=self.end,values_only=True): #获取指定行数据  2-7
            self.row_list = list(self.row)  # 将每一行的单元格值转换为列表
            self.new_row = self.row_list[0:self.d]
            self.lsz.append(self.new_row)
        return self.lsz
    def j_o_sum(self):
        for j in range(len(self.lsz)):
            for i in self.lsz[j][:]:
                if i % 2 == 0:
                    self.result1 += i
                else:
                    self.result2 += i
            self.osum.append(self.result1)
            self.jsum.append(self.result2)
        print('偶数和为:', self.osum)
        print('奇数和为:', self.jsum)
        for iz in range(self.first, self.end+1):
            self.ws[f'P{iz}'] = self.jsum[iz-2]
        for isz in range(self.first, self.end+1):
            self.ws[f'Q{isz}'] = self.osum[isz-2]
        self.wb.save(self.path)

cxy1 = cxy_add(r"C:\Users\19639\Desktop\求和.xlsx",2,7,15) #excel路径、获取2-7行数据、获取前15列数据
cxy1.creat_cxyexcel()
cxy1.j_o_sum()
