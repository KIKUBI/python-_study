# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   RONGHUA
# FileName:      day08.py
# Author:       陈啸宇
# Datetime:     2024/8/8 14:53
# Description:
# ---------------------------------------------------------------------------
import pymysql
import openpyxl
#获取一个连接对象
db =pymysql.connect(host="localhost", port=3306,user="root", password="123456", database="book")
#获取一个游标对象
cxy_db = db.cursor()
sql1 ='create table cxz ( \
        cx_id int primary key, \
        cx_name varchar(255), \
        cx_aderss varchar(255),\
        cx_phone varchar(255))'
sql2 = 'drop table cxz'
sql3 = "select * from cxz"
class cxy_excel:
    def __init__(self,cxy_db,db,path):
        self.cxy_db = cxy_db
        self.db = db
        self.w = openpyxl.load_workbook(path)
        self.ws = self.w.active
    def cxy_creat_table(self,sql):
        self.cxy_db.execute(sql)
        self.db.commit()
    def cxy_insertt_excel(self,path,first):
        for i in self.ws.iter_rows(min_row=first,values_only=True):
            sql = f"insert into cxz values ({i[0]},'{i[1]}','{i[2]}','{i[3]}')"
            self.cxy_db.execute(sql)
            db.commit()
        return self.ws
    def cxy_select_excel2(self,sql):
        self.cxy_db.execute(sql)
        self.date = self.cxy_db.fetchall()
        return self.date
    def cxy_append_excel(self,path,date):
        self.z = openpyxl.Workbook(path)
        self.zz = self.z.create_sheet('cxy',0)
        self.num = 1
        for i in date:
            self.li = list(i)
            self.li[0] = self.num
            self.zz.append(i)
            self.num+=1
        self.z.save(path)
cxy = cxy_excel(cxy_db,db,"cxy1.xlsx")
cxy.cxy_creat_table(sql1)
cxy.cxy_insertt_excel(2)
result = cxy.cxy_select_excel2(sql3)
cxy.cxy_append_excel("cxy.xlsx",result)

