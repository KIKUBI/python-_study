# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   test62
# FileName:      read_excel.py
# Author:       lao_zhao
# Datetime:     2024/8/14 15:52
# Description:
# 
# ---------------------------------------------------------------------------
import json

import openpyxl

from study_python.cxy_zuoye.sele2zuoye.common.read_ini import ReadIni


class ReadExcel:
    def __init__(self):
        """根据ReadIni的对象获取excel的路径，加载工作簿，获取工作表"""
        # 创建ReadIni对象
        ini = ReadIni()
        # 根据ReadIni的对象调用get_file_path获取excel的路径
        self.excel_path = ini.get_file_path("excel")
        # 根据ReadIni的对象调用get_table_name获取excel的工作表名称
        table_name = ini.get_table_name("tableName")
        # 加载工作簿
        self.wb = openpyxl.load_workbook(self.excel_path)
        # 获取工作表
        self.ws = self.wb[table_name]

    def get_data(self):
        """获取excel中测试需要使用的数据，并将这些属性存放在一个列表中"""
        # 创建一个列表存放每行的测试数据
        data_list = []
        # 循环excel中最大的行号
        for row in range(2, self.ws.max_row+1):
            # 获取每行中g列的用例数据, 并使用json模块中的loads将用例数据转成列表
            case_list = json.loads(self.ws["g"+str(row)].value)
            # 判断用例数据中最后一个元素的长度是否大于0，如果大于0将用例数据中的sql语句中？替换为组织编号
            if len(case_list[-1]) > 0:
                case_list[-1] = case_list[-1].replace("?", f'"{case_list[1]}"')

            # 获取h列的期望数据，并将期望数据存放到用例数据的列表中
            case_list.append(self.ws["h"+str(row)].value)
            # 将行号存放到用例数据的列表中
            case_list.append(row)
            # 将每行的数据存放到前面创建的列表中
            data_list.append(case_list)

        return data_list

    def write_result(self, row, result,  column="i"):
        self.ws[column+str(row)] = result
        self.wb.save(self.excel_path)


if __name__ == '__main__':
    excel = ReadExcel()
    print(excel.get_data())
    excel.write_result(2, "成功")