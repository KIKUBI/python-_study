# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   RONGHUA
# FileName:      excel.py
# Author:       陈啸宇
# Datetime:     2024/8/13 18:24
# Description:
# ---------------------------------------------------------------------------
import os

import openpyxl
from faker import Faker


class cxy_excel:
    def __init__(self):
        self.dates = Faker(locale="zh_CN")
        self.list1 = []
        if not os.path.exists('cxyz.xlsx'):
            self.excel = openpyxl.Workbook("cxy_xlsx")
            self.sheet = self.excel.create_sheet("cxy_用例",0)
            self.excel.save("cxyz.xlsx")
        else:
            self.excel = openpyxl.load_workbook("cxyz.xlsx")
            self.cxy = self.excel.create_sheet('cxy_用例z', 0)
            self.sheet = self.excel["cxy_用例"]
            self.insert_excel()
            self.insert_excelback()
            self.insert_excelback1()
            self.insert_excelback2()
            self.excel.save("cxyz.xlsx")
            for i in self.sheet.iter_rows(values_only=True):
                self.list1.append(i)
    def reday_excel(self):
        print(self.list1)
        return self.list1
    def insert_excel(self):
        for i in range(1,3):
            self.sheet[f'A{i}'] = self.dates.name()
            self.sheet[f'B{i}'] = self.dates.postcode()
            self.sheet[f'C{i}'] = f'DELETE FROM uc_org WHERE CODE_ = "{self.sheet[f"B{i}"].value}"'
            self.sheet[f'D{i}'] = '创建成功'
    def insert_excelback(self):
        for i in range(3,5):
            self.sheet[f'A{i}'] = ''
            self.sheet[f'B{i}'] = self.dates.postcode()
            self.sheet[f'C{i}'] = ''
            self.sheet[f'D{i}'] = '创建失败'
    def insert_excelback1(self):
        for i in range(5,7):
            self.sheet[f'A{i}'] = self.dates.name()
            self.sheet[f'B{i}'] = ''
            self.sheet[f'C{i}'] = ''
            self.sheet[f'D{i}'] = '创建失败'
    def insert_excelback2(self):
        for i in range(7,9):
            self.sheet[f'A{i}'] = ''
            self.sheet[f'B{i}'] = ''
            self.sheet[f'C{i}'] = ''
            self.sheet[f'D{i}'] = '创建失败'
    def write_excel(self,result,z):
        self.li = list(z)
        print(self.li)
        self.li.append(result)
        self.cxy.append(self.li)
        self.excel.save("cxyz.xlsx")
cxyexcl = cxy_excel()




