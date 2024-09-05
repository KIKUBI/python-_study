# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   test62
# FileName:      eg_2.py
# Author:       lao_zhao
# Datetime:     2024/8/8 9:48
# Description:
# 
# ---------------------------------------------------------------------------
import os

import openpyxl

'''
# 1:加载工作簿
wb1 = openpyxl.load_workbook("./file/用户表1.xlsx")
wb2 = openpyxl.load_workbook("./file/用户表2.xlsx")
# 2:获取工作表
ws1 = wb1["数据"]
ws2 = wb2["数据"]
# 3:获取工作表中的数据
# 3.1 先获取第一个表的数据
# cell = ws["列号行号"].value
# 创建一个列表，用于存放工作表中的数据
sheet_list1 = []
# 直接循环工作表
for row in ws1:
    # 创建列表，用于存放每行的内容
    row_list = []
    # 循环工作表之后，获取到每行的元组
    # 循环行，得到每行的单元格
    for cell in row:
        # 将每行中单元格数据追加到列表中
        row_list.append(cell.value)

    # 将存放每行数据的列表追加到前面创建大列表中
    sheet_list1.append(row_list)

# 3.2 再获取第二个表的数据
# cell = ws["列号行号"].value
# 创建一个列表，用于存放工作表中的数据
sheet_list2 = []
# 直接循环工作表
for row in ws1:
    # 创建列表，用于存放每行的内容
    row_list = []
    # 循环工作表之后，获取到每行的元组
    # 循环行，得到每行的单元格
    for cell in row:
        # 将每行中单元格数据追加到列表中
        row_list.append(cell.value)

    # 将存放每行数据的列表追加到前面创建大列表中
    sheet_list2.append(row_list)


# 两个工作表的数据进行相加
# 判断第二章表数据中是否有表头，如果有去掉，如果每个就不用去掉
if ['编号', '姓名', '地址', '所在公司', '手机号'] in sheet_list2:
    sum_list = sheet_list1 + sheet_list2[1:]
else:
    sum_list = sheet_list1 + sheet_list2


# 创建工作簿
new_wb = openpyxl.Workbook()
# 创建工作表
new_ws = new_wb.create_sheet("数据汇总", 0)

# 循环合并之后的数据，处理编号
number = 1
for i in sum_list:
    # 判断i的第一个元素是否为整形如果是，就替换第一个元素的值，如果不是整形，表示就是表头
    if isinstance(i[0], int):
        i[0] = number
        number += 1
    # 处理之后的数据写入到新的工作表中
    new_ws.append(i)

# 保存创建的工作簿
new_wb.save("./file/用户汇总表.xlsx")
'''


def get_table_data(excel_path, table_name):
    # 1:加载工作簿
    wb1 = openpyxl.load_workbook(excel_path)
    # 2:获取工作表
    ws1 = wb1[table_name]
    # 3:获取工作表中的数据
    # 3.1 先获取第一个表的数据
    # cell = ws["列号行号"].value
    # 创建一个列表，用于存放工作表中的数据
    sheet_list1 = []
    # 直接循环工作表
    for row in ws1:
        # 创建列表，用于存放每行的内容
        row_list = []
        # 循环工作表之后，获取到每行的元组
        # 循环行，得到每行的单元格
        for cell in row:
            # 将每行中单元格数据追加到列表中
            row_list.append(cell.value)

        # 将存放每行数据的列表追加到前面创建大列表中
        sheet_list1.append(row_list)
    else:
        return sheet_list1


def write_table_data(*datas, excel_path=None, table_name=None):
    # 创建工作簿
    new_wb = openpyxl.Workbook()
    # 创建工作表
    new_ws = new_wb.create_sheet(table_name, 0)

    # 两个工作表的数据进行相加

    # 创建一个列表，存放所有列表中的元素，这个列表中只存放表头使用
    sum_list = [['编号', '姓名', '地址', '所在公司', '手机号']]
    # 判断datas的长度是否大于1
    if len(datas) > 0:
        for data in datas:
            # 判断第二章表数据中是否有表头，如果有去掉，如果每个就不用去掉
            if ['编号', '姓名', '地址', '所在公司', '手机号'] in data:
                sum_list = sum_list + data[1:]
            else:
                sum_list = sum_list + data

    if len(sum_list) > 2:
        # 循环合并之后的数据，处理编号
        number = 1
        for i in sum_list:
            # 判断i的第一个元素是否为整形如果是，就替换第一个元素的值，如果不是整形，表示就是表头
            if isinstance(i[0], int):
                i[0] = number
                number += 1
            # 处理之后的数据写入到新的工作表中
            new_ws.append(i)

        # 保存创建的工作簿
        new_wb.save(excel_path)


class ExcelFunc:
    def __init__(self, excel_path, table_name):
        """创建或加载工作簿，根据工作表创建或获取工作表"""
        # 将excel的路径创建为对象属性
        self.excel_path = excel_path
        # 判断excel的路径是否存在并且是.xlsx结尾
        if os.path.isfile(self.excel_path) and self.excel_path.endswith(".xlsx"):
            # 加载工作簿
            self.wb = openpyxl.load_workbook(self.excel_path)
            # 根据工作簿获取工作表
            self.ws = self.wb[table_name]
        # 判断excel的路径是不存在但是是.xlsx结尾
        elif not os.path.isfile(self.excel_path) and self.excel_path.endswith(".xlsx"):
            # 创建工作簿
            self.wb = openpyxl.Workbook()
            # 根据工作簿创建工作表
            self.ws = self.wb.create_sheet(table_name, 0)
        # excel的路径不存在，也不是.xlsx结尾
        else:
            print("请求正确传入excel的路径")

    def get_table_data(self):
        # # 1:加载工作簿
        # wb1 = openpyxl.load_workbook(excel_path)
        # # 2:获取工作表
        # ws1 = wb1[table_name]
        # 3:获取工作表中的数据
        # 3.1 先获取第一个表的数据
        # cell = ws["列号行号"].value
        # 创建一个列表，用于存放工作表中的数据
        sheet_list1 = []
        # 直接循环工作表
        for row in self.ws:
            # 创建列表，用于存放每行的内容
            row_list = []
            # 循环工作表之后，获取到每行的元组
            # 循环行，得到每行的单元格
            for cell in row:
                # 将每行中单元格数据追加到列表中
                row_list.append(cell.value)

            # 将存放每行数据的列表追加到前面创建大列表中
            sheet_list1.append(row_list)
        else:
            return sheet_list1

    def write_table_data(self, *datas, title=['编号', '姓名', '地址', '所在公司', '手机号']):
        # # 创建工作簿
        # new_wb = openpyxl.Workbook()
        # # 创建工作表
        # new_ws = new_wb.create_sheet(table_name, 0)

        # 两个工作表的数据进行相加

        # 创建一个列表，存放所有列表中的元素，这个列表中只存放表头使用
        sum_list = [title]
        # 判断datas的长度是否大于1
        if len(datas) > 0:
            for data in datas:
                # 判断第二章表数据中是否有表头，如果有去掉，如果每个就不用去掉
                if title in data:
                    sum_list = sum_list + data[1:]
                else:
                    sum_list = sum_list + data

        if len(sum_list) > 2:
            # 循环合并之后的数据，处理编号
            number = 1
            for i in sum_list:
                # 将i转成列表
                i = list(i)
                # 判断i的第一个元素是否为整形如果是，就替换第一个元素的值，如果不是整形，表示就是表头
                if isinstance(i[0], int):
                    i[0] = number
                    number += 1
                # 处理之后的数据写入到新的工作表中
                self.ws.append(i)

            # 保存创建的工作簿
            self.wb.save(self.excel_path)


if __name__ == '__main__':
    # datas1 = get_table_data("./file/用户表1.xlsx", "数据")
    #
    # datas2 = get_table_data("./file/用户表2.xlsx", "数据")
    #
    # write_table_data(datas1, datas2, excel_path="./file/用户表3.xlsx", table_name="数据汇总")

    excel1 = ExcelFunc("./file/用户表1.xlsx", "数据")
    datas1 = excel1.get_table_data()

    excel2 = ExcelFunc("./file/用户表2.xlsx", "数据")
    datas2 = excel1.get_table_data()

    excel3 = ExcelFunc("./file/用户表.xlsx", "数据")
    excel3.write_table_data(datas1, datas2)