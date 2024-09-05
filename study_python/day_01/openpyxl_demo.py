# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   test62
# FileName:      openpyxl_demo.py
# Author:       lao_zhao
# Datetime:     2024/8/7 16:00
# Description:
# 
# ---------------------------------------------------------------------------
import datetime
import time
from collections.abc import Iterable

import openpyxl


"""
    创建流程：
    1: 创建工作簿
        工作簿 = openpyxl.Workbook()

    2：使用工作簿创建工作表
        工作表 = 工作簿.create_sheet("工作表的名称"[, number])
        工作表 = 工作簿.active   # 使用工作簿自带的工作表
        
    3：向工作表写入数据
        工作表["列号行号"] = 值
        工作表.append(有序的可迭代类型)  注意：有序的可迭代类型=(列表，元组，range序列， 字典-字典的key必须为列号)
            将有序的可迭代类型的元素写入到没有数据的新行，一个元素占一个格子。
        
    4：保存工作簿
        工作簿.save("excel文件的路径")
"""
"""
    读取流程：
    1：加载工作簿
        工作簿 = openpyxl.load_workbook("excel路径")
        
    2：获取工作表
        工作表 = 工作簿["工作表的名称"]
            注意：工作表是可迭代类型
            
    3：读取工作表中的数据
        i：根据单元格获取数据
            单元格 = 工作表["列号行号"]
            单元格的操作：
                获取单元格数据： 单元格.value
                获取单元格行号： 单元格.row
                获取单元格列号： 单元格.column
                获取单元格坐标： 单元格.coordinate
        2：根据工作表获取数据-因为工作表是可迭代类型，可以直接循环工作表，
            会获取工作表中有数据的所有的行,行的类型为元组，元组中存放每个单元格。

"""

'''
wb = openpyxl.Workbook()

ws = wb.create_sheet("我的工作表", 0)
# ws = wb.active
ws["b1"] = 100
ws["b2"] = "成都"
ws["C2"] = datetime.datetime.now()
ws["C3"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
ws["C4"] = None
ws["C5"] = True
# ws["C6"] = {"True": 1, "False": 2}

ws.append([1, 2, 3, 4])
ws.append((1, 2, 3, 4))
ws.append(range(10))
ws.append({"a": 100, "B": 1, "c": 123})

wb.save("./file/data.xlsx")
'''


wb = openpyxl.load_workbook("./file/data.xlsx")
ws = wb["我的工作表"]

print(ws, type(ws))

print(isinstance(ws, Iterable))

cell = ws["b1"]
# print("单元格数据", cell.value)
# print("获取单元格行号", cell.row)
# print("获取单元格列号", cell.column)
# print("获取单元格坐标", cell.coordinate)
print(cell, type(cell))


print(ws["b2"].value)


for row in ws:
    # print(row, type(row))
    for cell in row:
        print([cell.value, type(cell.value)], end="\t")
    print()


ws["D1"] = "你好"
wb.save("./file/data.xlsx")