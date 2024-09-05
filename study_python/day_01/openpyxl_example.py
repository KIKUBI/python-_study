# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   test62
# FileName:      openpyxl_example.py
# Author:       lao_zhao
# Datetime:     2024/8/7 17:03
# Description:
# 
# ---------------------------------------------------------------------------
import openpyxl
from faker import Faker

'''
wb = openpyxl.Workbook()

ws = wb.create_sheet("数据", 0)

ws.append(["编号", '姓名', "地址", "所在公司", "手机号"])


data = Faker(locale="zh_cn")

for i in range(1, 301):
    row_data = [i, data.name(), data.address(), data.company(), data.phone_number()]
    ws.append(row_data)

wb.save("./file/用户表1.xlsx")


# 复制excel
with open("./file/用户表1.xlsx", mode="rb") as f1:
    content = f1.read()

with open("./file/用户表2.xlsx", mode="wb") as fp:
    fp.write(content)
'''


wb = openpyxl.load_workbook("D:\桌面\求和.xlsx")

ws = wb["Sheet1"]

for row in ws:
    row_list = {}
    for cell in row:
        row_list[cell.coordinate] = cell.value

    print(list(row_list.values()))
    print(list(row_list.keys()))